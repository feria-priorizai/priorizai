import csv
import io
import math
from collections.abc import AsyncIterator, Iterable, Sequence
from contextlib import asynccontextmanager
from dataclasses import dataclass, field
from datetime import UTC, datetime
from uuid import uuid4

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from sqlalchemy import inspect, select, text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.api.interconsultas import router as interconsultas_router
from app.core.config import settings
from app.core.database import SessionLocal, engine
from app.models import Base, Interconsulta
from app.services.banderas_rojas import aplicar_banderas_a_interconsulta
from app.services.ner import get_extractor_ner
from app.services.priorizador import (
    aplicar_resultado,
    get_priorizador,
    tiene_informacion_clinica,
)

UPLOAD_FILE = File(...)
CAMPOS_OBLIGATORIOS_FORM = Form(None)

COLUMNAS_ESPERADAS = [
    "ESPEC_ORIGEN",
    "EDAD",
    "SEXO",
    "ESPEC_DESTINO",
    "HISTORIA_CLINICA",
    "FUNDAMENTOS_DIAGNOSTICO",
    "EXAMENES_COMPLEMENTARIOS",
    "MOTIVO_INTERCONSULTA",
]

# PRIORIDAD no es obligatoria: en produccion la interconsulta llega SIN priorizar
# (ese es el producto). Solo la traen los archivos historicos, donde es la
# etiqueta que asigno un especialista. Se guarda cuando viene, para poder
# contrastar despues el modelo contra la prioridad real, pero no se exige ni se
# muestra como si fuera la prioridad de la interconsulta.
COLUMNA_PRIORIDAD_OPCIONAL = "PRIORIDAD"

# Campos que DEBEN tener valor en CADA fila del archivo. Debe coincidir con
# los campos marcados como obligatorios por defecto en la pestaña de
# configuración (obligatorioPorDefecto: true).
COLUMNAS_OBLIGATORIAS_POR_FILA = [
    "ESPEC_ORIGEN",
    "EDAD",
    "SEXO",
    "ESPEC_DESTINO",
    "HISTORIA_CLINICA",
    "FUNDAMENTOS_DIAGNOSTICO",
    "MOTIVO_INTERCONSULTA",
]

# EDAD no se puede volver opcional desde la configuracion: la columna es NOT
# NULL y, a diferencia de los campos de texto, no tiene un vacio que guardar.
COLUMNAS_SIEMPRE_OBLIGATORIAS = ["EDAD"]

# Lo unico que la configuracion puede marcar como obligatorio. Las claves de la
# pestana de configuracion incluyen campos que no vienen en el archivo
# (PRIORIDAD_ACTUAL, ID, ESTADO...); esas se ignoran en vez de rechazarse.
COLUMNAS_CONFIGURABLES = frozenset(COLUMNAS_ESPERADAS)


@asynccontextmanager
async def lifespan(_: FastAPI) -> AsyncIterator[None]:
    """Toda la conversacion con la base ocurre aca, no al importar el modulo.

    Con `create_all` y `_asegurar_columnas_interconsultas()` a nivel de modulo,
    `import app.main` abria conexion: los tests no corrian sin Postgres y la CI
    levantaba un servicio entero solo para poder importar la app."""
    settings.verificar_credenciales()
    Base.metadata.create_all(bind=engine)
    _asegurar_columnas_interconsultas()
    yield


app = FastAPI(lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    # El listado publica el total real ahi; sin exponerla, el navegador la oculta.
    expose_headers=["X-Total-Count"],
)

app.include_router(interconsultas_router)


@app.get("/")
def root() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "healthy"}


@app.post("/upload-csv")
async def upload_csv(
    file: UploadFile = UPLOAD_FILE,
    campos_obligatorios: str | None = CAMPOS_OBLIGATORIOS_FORM,
) -> dict[str, object]:
    """`campos_obligatorios` viene de la pestana de configuracion, separado por
    comas. Sin el se usan los de fabrica: antes la configuracion no salia del
    navegador y la pestana no tenia ningun efecto sobre la carga."""
    filename = (file.filename or "").lower()
    es_csv = filename.endswith(".csv")
    es_xlsx = filename.endswith(".xlsx")
    if not (es_csv or es_xlsx):
        raise HTTPException(
            status_code=400,
            detail="El archivo debe ser CSV o XLSX valido",
        )

    contenido = await file.read()
    if es_xlsx:
        filas = _leer_xlsx(contenido)
        tipo_archivo = "xlsx"
    else:
        filas = _leer_csv(contenido)
        tipo_archivo = "csv"

    resultado_validacion = _validar_filas(
        filas, _resolver_obligatorias(campos_obligatorios)
    )
    guardado = _guardar_interconsultas(
        resultado_validacion["filas_validas"], tipo_archivo
    )
    return {
        **guardado,
        "rejected": resultado_validacion["rejected"],
        "rejected_count": len(resultado_validacion["rejected"]),
    }


@dataclass(frozen=True)
class FilaLeida:
    """Una fila del archivo con su numero real de linea.

    El numero se arrastra desde la lectura porque las filas en blanco y las mal
    formadas no llegan a la validacion: numerarlas alli corria el indice y el
    modal de errores senalaba una fila distinta a la del archivo.
    """

    numero: int
    datos: dict[str, str] = field(default_factory=dict)
    # Problema estructural (columnas de mas o de menos). La fila se rechaza sin
    # detener el resto del archivo.
    error: str | None = None


def _leer_csv(contenido: bytes) -> list[FilaLeida]:
    try:
        texto = contenido.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise HTTPException(
            status_code=400,
            detail="El CSV debe estar codificado en UTF-8",
        ) from None

    try:
        dialect = csv.Sniffer().sniff(texto[:4096], delimiters=",;")
    except csv.Error:
        dialect = csv.excel

    try:
        reader_rows = list(csv.reader(io.StringIO(texto), dialect=dialect))
    except csv.Error as error:
        raise HTTPException(
            status_code=400,
            detail=f"Error al parsear el CSV: {error}",
        ) from error

    if not reader_rows:
        raise HTTPException(status_code=400, detail="El CSV no contiene filas")

    header = [_normalizar_valor(cell) for cell in reader_rows[0]]
    filas = []
    for numero, row in enumerate(reader_rows[1:], start=2):
        valores = [_normalizar_valor(cell) for cell in row]
        if not valores or all(not cell for cell in valores):
            continue
        if len(valores) != len(header):
            filas.append(
                FilaLeida(
                    numero=numero,
                    datos=dict(zip(header, valores, strict=False)),
                    error=(
                        f"Se esperaban {len(header)} columnas y "
                        f"llegaron {len(valores)}"
                    ),
                )
            )
            continue
        filas.append(
            FilaLeida(numero=numero, datos=dict(zip(header, valores, strict=True)))
        )
    return filas


def _leer_xlsx(contenido: bytes) -> list[FilaLeida]:
    try:
        workbook = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception as error:
        raise HTTPException(
            status_code=400,
            detail=f"No se pudo leer el XLSX: {error}",
        ) from error

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="El XLSX no contiene filas")

    header = [_normalizar_valor(cell) for cell in rows[0]]
    filas = []
    for numero, row in enumerate(rows[1:], start=2):
        valores = [_normalizar_valor(cell) for cell in row[: len(header)]]
        if not valores or all(not cell for cell in valores):
            continue
        extra = row[len(header) :]
        if any(_normalizar_valor(cell) for cell in extra):
            filas.append(
                FilaLeida(
                    numero=numero,
                    datos=dict(zip(header, valores, strict=False)),
                    error="Contiene columnas extra con datos",
                )
            )
            continue
        filas.append(
            FilaLeida(numero=numero, datos=dict(zip(header, valores, strict=True)))
        )
    return filas


def _resolver_obligatorias(campos: str | None) -> list[str]:
    """Campos que deben venir con valor, segun la configuracion del cliente.

    Las claves que no son columnas del archivo (PRIORIDAD_ACTUAL, ID, ESTADO...)
    se ignoran: la pestana de configuracion las ofrece para exportar.
    """
    if campos is None:
        return list(COLUMNAS_OBLIGATORIAS_POR_FILA)

    pedidos = [
        _normalizar_encabezado(campo) for campo in campos.split(",") if campo.strip()
    ]
    elegidos = [campo for campo in pedidos if campo in COLUMNAS_CONFIGURABLES]
    for campo in COLUMNAS_SIEMPRE_OBLIGATORIAS:
        if campo not in elegidos:
            elegidos.append(campo)
    return elegidos


def _validar_filas(
    filas: Sequence[FilaLeida],
    obligatorias: Iterable[str] | None = None,
) -> dict[str, list]:
    """Separa las filas utilizables de las rechazadas.

    Ninguna fila mala detiene el archivo: se junta el motivo de cada una en
    `rejected` para que el cliente pueda mostrarlas.
    """
    if not filas:
        raise HTTPException(
            status_code=400,
            detail="El archivo no contiene filas de datos",
        )

    requeridas = (
        list(COLUMNAS_OBLIGATORIAS_POR_FILA)
        if obligatorias is None
        else list(obligatorias)
    )

    encabezados = [_normalizar_encabezado(h) for h in filas[0].datos]
    faltantes_enc = [h for h in requeridas if h not in encabezados]
    if faltantes_enc:
        unidos = ", ".join(faltantes_enc)
        raise HTTPException(
            status_code=400,
            detail=f"Faltan encabezados obligatorios: {unidos}",
        )

    filas_validas: list[dict[str, object]] = []
    rejected: list[dict[str, object]] = []

    for fila in filas:
        datos: dict[str, object] = {
            _normalizar_encabezado(clave): _normalizar_valor(valor)
            for clave, valor in fila.datos.items()
            if clave is not None
        }

        if fila.error is not None:
            _rechazar(rejected, fila.numero, fila.error, datos)
            continue

        faltantes = [
            campo for campo in requeridas if not str(datos.get(campo, "")).strip()
        ]
        if faltantes:
            _rechazar(rejected, fila.numero, faltantes, datos)
            continue

        edad = _parsear_edad(datos.get("EDAD"))
        if edad is None:
            _rechazar(rejected, fila.numero, "EDAD (formato inválido)", datos)
            continue
        if not 0 <= edad <= settings.edad_maxima:
            _rechazar(rejected, fila.numero, "EDAD (fuera de rango)", datos)
            continue
        datos["EDAD"] = edad

        filas_validas.append(datos)

    return {"filas_validas": filas_validas, "rejected": rejected}


def _rechazar(
    rejected: list[dict[str, object]],
    numero: int,
    motivo: str | list[str],
    datos: dict[str, object],
) -> None:
    rejected.append(
        {
            "fila": numero,
            "campos_faltantes": [motivo] if isinstance(motivo, str) else motivo,
            "datos_raw": datos,
        }
    )


def _guardar_interconsultas(
    filas_json: list[dict[str, object]], tipo_archivo: str
) -> dict[str, object]:
    session = SessionLocal()
    try:
        insert_sql = text("""
            INSERT INTO interconsultas
            (id, espec_origen, edad, sexo, espec_destino, prioridad_original_csv,
             historia_clinica, fundamentos_diagnostico, examenes_complementarios,
             motivo_interconsulta, fecha_emision, estado, bandera_roja,
             prioridad_forzada_por_regla, created_at, updated_at)
            VALUES
            (:id, :espec_origen, :edad, :sexo, :espec_destino, :prioridad_original_csv,
             :historia_clinica, :fundamentos_diagnostico, :examenes_complementarios,
             :motivo_interconsulta, :fecha_emision, :estado, :bandera_roja,
             :prioridad_forzada_por_regla, :created_at, :updated_at)
            """)

        ids_insertados: list[str] = []
        for fila_json in filas_json:
            interconsulta_id = str(uuid4())
            ahora = datetime.now(UTC)
            params: dict[str, object | None] = {
                "id": interconsulta_id,
                "espec_origen": fila_json.get("ESPEC_ORIGEN", ""),
                "edad": fila_json.get("EDAD"),
                "sexo": fila_json.get("SEXO", ""),
                "espec_destino": fila_json.get("ESPEC_DESTINO", ""),
                "prioridad_original_csv": _texto_o_none(
                    fila_json.get(COLUMNA_PRIORIDAD_OPCIONAL)
                ),
                "historia_clinica": fila_json.get("HISTORIA_CLINICA", ""),
                "fundamentos_diagnostico": fila_json.get("FUNDAMENTOS_DIAGNOSTICO", ""),
                "examenes_complementarios": fila_json.get(
                    "EXAMENES_COMPLEMENTARIOS", ""
                ),
                "motivo_interconsulta": fila_json.get("MOTIVO_INTERCONSULTA", ""),
                "fecha_emision": _parsear_fecha_emision(fila_json.get("FECHA_EMISION")),
                "estado": "pendiente",
                "bandera_roja": False,
                "prioridad_forzada_por_regla": False,
                "created_at": ahora,
                "updated_at": ahora,
            }
            session.execute(insert_sql, params)
            ids_insertados.append(interconsulta_id)

        priorizadas = _priorizar_interconsultas_insertadas(session, ids_insertados)
        session.commit()
        return {
            "inserted": len(filas_json),
            "stored": len(filas_json),
            "file_type": tipo_archivo,
            "prioritized": priorizadas,
            "prioritization_status": _estado_priorizacion(
                total=len(filas_json),
                priorizadas=priorizadas,
            ),
            "ids": ids_insertados,
        }
    except HTTPException:
        session.rollback()
        raise
    except SQLAlchemyError as error:
        session.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Error en la base de datos: {error}",
        ) from error
    finally:
        session.close()


COLUMNAS_NUEVAS = {
    "estado": "VARCHAR(20) DEFAULT 'pendiente' NOT NULL",
    "motivo_sin_prioridad": "TEXT",
    "fecha_emision": "TIMESTAMP",
    "bandera_roja": "BOOLEAN DEFAULT false NOT NULL",
    "terminos_bandera_roja": "TEXT",
    "prioridad_forzada_por_regla": "BOOLEAN DEFAULT false NOT NULL",
    "entidades": "JSON",
    "entidades_error": "TEXT",
}


def _asegurar_columnas_interconsultas() -> None:
    inspector = inspect(engine)
    columnas = {columna["name"] for columna in inspector.get_columns("interconsultas")}
    faltantes = {
        nombre: definicion
        for nombre, definicion in COLUMNAS_NUEVAS.items()
        if nombre not in columnas
    }
    if not faltantes:
        return

    with engine.begin() as connection:
        for nombre, definicion in faltantes.items():
            connection.execute(
                text(f"ALTER TABLE interconsultas ADD COLUMN {nombre} {definicion}")
            )


def _priorizar_interconsultas_insertadas(session: Session, ids: list[str]) -> int:
    interconsultas = list(
        session.scalars(
            select(Interconsulta).where(Interconsulta.id.in_(ids)),
        ).all(),
    )
    validas = [
        interconsulta
        for interconsulta in interconsultas
        if tiene_informacion_clinica(interconsulta)
    ]

    for interconsulta in validas:
        aplicar_banderas_a_interconsulta(interconsulta, ya_modificada_por_medico=False)

    _extraer_entidades(validas)

    try:
        resultados = get_priorizador().predecir(validas)
    except Exception as exc:
        for interconsulta in validas:
            interconsulta.motivo_sin_prioridad = (
                f"No se pudo ejecutar el modelo predictivo: {exc}"
            )
        return 0

    por_id = {interconsulta.id: interconsulta for interconsulta in validas}
    for resultado in resultados:
        aplicar_resultado(por_id[resultado.id], resultado)

    return len(resultados)


def _extraer_entidades(interconsultas: list[Interconsulta]) -> int:
    """Corre el NER sobre los campos clinicos y guarda las entidades.

    Un fallo del modelo no debe tumbar la carga: se registra en
    entidades_error y la interconsulta se guarda igual, como con el
    priorizador.
    """
    if not interconsultas:
        return 0

    try:
        extractor = get_extractor_ner()
    except Exception as exc:
        for interconsulta in interconsultas:
            interconsulta.entidades_error = f"No se pudo cargar el modelo NER: {exc}"
        return 0

    procesadas = 0
    for interconsulta in interconsultas:
        try:
            interconsulta.entidades = extractor.extraer_de_interconsulta(interconsulta)
            interconsulta.entidades_error = None
            procesadas += 1
        except Exception as exc:
            interconsulta.entidades = None
            interconsulta.entidades_error = f"No se pudo extraer entidades: {exc}"
    return procesadas


def _estado_priorizacion(total: int, priorizadas: int) -> str:
    if total == 0:
        # Todas las filas se rechazaron: no hubo priorizacion que completar.
        return "skipped"
    if priorizadas == total:
        return "completed"
    if priorizadas == 0:
        return "skipped"
    return "partial"


def _texto_o_none(valor: object) -> str | None:
    """Distingue 'no vino la columna' de 'vino vacia': ambos casos se guardan como
    NULL, para que coalesce y las comparaciones no tengan que lidiar con ""."""
    texto = str(valor or "").strip()
    return texto or None


FORMATOS_FECHA_EMISION = ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y")


def _parsear_fecha_emision(valor: object) -> datetime | None:
    """FECHA_EMISION es opcional (HU3-c1): no viene en COLUMNAS_ESPERADAS y aun no
    se conoce el formato real del archivo del sistema hospitalario (D18). Si no
    esta presente o no calza con un formato conocido, se guarda como None en vez
    de romper la carga completa.

    Se guarda como datetime naive a proposito: es una fecha de calendario, no un
    instante. Marcarla como UTC hacia que el frontend la convirtiera a horario de
    Chile y mostrara el dia anterior.
    """
    texto = str(valor or "").strip()
    if not texto:
        return None

    for formato in FORMATOS_FECHA_EMISION:
        try:
            return datetime.strptime(texto, formato)
        except ValueError:
            continue
    return None


def _parsear_edad(valor: object) -> int | None:
    """Anios cumplidos a partir de la celda del archivo.

    Un CSV entrega cada celda como texto, asi que "53.0" -lo que exporta
    cualquier planilla- llegaba como cadena y borrarle el punto lo convertia
    en 530. El separador se interpreta como decimal, nunca como separador de
    miles: en una edad no tiene sentido. Los decimales se truncan, que es la
    convencion clinica (4,5 anios son 4 anios cumplidos).

    Los espacios internos no se ignoran: "4 6" es mas probablemente un error
    de tipeo que una edad de 46 anios.

    Devuelve None si el valor no es un numero; el rango lo valida la llamante,
    para poder distinguir 'no es un numero' de 'no es una edad posible'.
    """
    if valor is None:
        return None

    texto = str(valor).strip()
    if not texto:
        return None

    normalizado = texto.replace(",", ".")
    if normalizado.count(".") > 1:
        return None

    try:
        return math.floor(float(normalizado))
    except (ValueError, OverflowError):
        return None


def _normalizar_encabezado(valor: object) -> str:
    return _normalizar_valor(valor).upper()


def _normalizar_valor(valor: object) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return str(valor).replace("\ufeff", "").replace("\xa0", " ").strip()
