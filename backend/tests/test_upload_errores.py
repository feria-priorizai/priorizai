"""Ramas de error de la ingesta (`POST /upload-csv`).

Cubre los rechazos de formato de `_leer_csv`, `_leer_xlsx` y `_validar_filas`, y
el rollback ante un fallo de base de datos.
"""

import csv
from collections.abc import Iterator, Sequence
from io import BytesIO

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy.exc import SQLAlchemyError

import app.main as main_module

HEADER = (
    "ESPEC_ORIGEN,EDAD,SEXO,ESPEC_DESTINO,PRIORIDAD,HISTORIA_CLINICA,"
    "FUNDAMENTOS_DIAGNOSTICO,EXAMENES_COMPLEMENTARIOS,MOTIVO_INTERCONSULTA\n"
)
FILA = (
    "MEDICINA GENERAL,46,FEMENINO,RESPIRATORIO ADULTO,ALTA,"
    "CANCER PULMONAR,Paciente estable.,,CONTROL DE ESPECIALIDAD\n"
)


# Una linea en blanco y una fila de solo comas: las dos se saltean al leer.
BLANCO = "\n"
VACIA = ",,,,,,,,\n"
ENCABEZADOS_XLSX = [
    "ESPEC_ORIGEN",
    "EDAD",
    "SEXO",
    "ESPEC_DESTINO",
    "PRIORIDAD",
    "HISTORIA_CLINICA",
    "FUNDAMENTOS_DIAGNOSTICO",
    "EXAMENES_COMPLEMENTARIOS",
    "MOTIVO_INTERCONSULTA",
]
FILA_XLSX = [
    "MEDICINA GENERAL",
    46,
    "FEMENINO",
    "RESPIRATORIO ADULTO",
    "ALTA",
    "CANCER PULMONAR",
    "Paciente estable.",
    "",
    "CONTROL DE ESPECIALIDAD",
]

# Fila con menos columnas que el encabezado.
CORTA = "MEDICINA GENERAL,46\n"


def _subir(client: TestClient, nombre: str, contenido: bytes | str):
    return client.post("/upload-csv", files={"file": (nombre, contenido, "text/csv")})


# --------------------------------------------------------------------------
# CSV: formato del archivo
# --------------------------------------------------------------------------


def test_csv_no_codificado_en_utf8_devuelve_400(client: TestClient) -> None:
    contenido = (HEADER + FILA).encode("latin-1").replace(b"CANCER", b"CANCI\xd3N")

    response = _subir(client, "latin1.csv", contenido)

    assert response.status_code == 400
    assert response.json()["detail"] == "El CSV debe estar codificado en UTF-8"


def test_csv_vacio_devuelve_400(client: TestClient) -> None:
    response = _subir(client, "vacio.csv", b"")

    assert response.status_code == 400
    assert response.json()["detail"] == "El CSV no contiene filas"


@pytest.fixture
def limite_de_campo_csv() -> Iterator[None]:
    """Fuerza el `csv.Error` de `_leer_csv` sin depender de bytes raros."""
    previo = csv.field_size_limit()
    csv.field_size_limit(32)
    try:
        yield
    finally:
        csv.field_size_limit(previo)


def test_csv_ilegible_devuelve_400(
    client: TestClient,
    limite_de_campo_csv: None,
) -> None:
    contenido = HEADER + "MEDICINA GENERAL," + "x" * 500 + ",F,CARDIO,ALTA,H,F,,M\n"

    response = _subir(client, "corrupto.csv", contenido)

    assert response.status_code == 400
    assert "Error al parsear el CSV" in response.json()["detail"]


def test_csv_con_columnas_de_mas_rechaza_solo_esa_fila(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    """Regresion: cualquier fila mal formada devolvia 400 y no entraba
    ninguna, aunque el resto del archivo estuviera bien."""
    contenido = HEADER + FILA + FILA.replace("\n", ",columna extra\n")

    response = _subir(client, "extra.csv", contenido)

    assert response.status_code == 200
    body = response.json()
    assert body["inserted"] == 1
    assert body["rejected_count"] == 1
    assert body["rejected"][0]["fila"] == 3


def test_csv_saltea_las_filas_completamente_vacias(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    contenido = HEADER + FILA + "\n" + ",,,,,,,,\n" + FILA

    response = _subir(client, "con_vacias.csv", contenido)

    assert response.status_code == 200
    body = response.json()
    assert body["inserted"] == 2
    assert body["rejected_count"] == 0


# --------------------------------------------------------------------------
# XLSX: formato del archivo
# --------------------------------------------------------------------------


def _xlsx(filas: Sequence[Sequence[object]]) -> bytes:
    workbook = Workbook()
    hoja = workbook.active
    for fila in filas:
        hoja.append(fila)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_xlsx_ilegible_devuelve_400(client: TestClient) -> None:
    response = _subir(client, "roto.xlsx", b"esto no es un xlsx")

    assert response.status_code == 400
    assert "No se pudo leer el XLSX" in response.json()["detail"]


def test_xlsx_sin_filas_devuelve_400(client: TestClient) -> None:
    response = _subir(client, "vacio.xlsx", _xlsx([]))

    assert response.status_code == 400
    assert response.json()["detail"] == "El XLSX no contiene filas"


def test_xlsx_con_columnas_extra_absorbe_la_columna_sin_nombre(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    # openpyxl rellena TODAS las filas hasta el ancho de la hoja, encabezado
    # incluido. Por eso `len(header)` nunca es menor que `len(row)` y el guard de
    # "columnas extra con datos" (`main.py:172-177`) no se alcanza por esta via:
    # la columna de mas entra como un encabezado vacio y su dato se descarta.
    encabezados = HEADER.strip().split(",")
    fila = [
        "MEDICINA GENERAL",
        46,
        "FEMENINO",
        "RESPIRATORIO ADULTO",
        "ALTA",
        "CANCER PULMONAR",
        "Paciente estable",
        "",
        "CONTROL",
        "dato de mas",
    ]

    response = _subir(client, "extra.xlsx", _xlsx([encabezados, fila]))

    assert response.status_code == 200
    assert response.json()["inserted"] == 1


def test_xlsx_acepta_la_edad_como_float(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    # openpyxl entrega los enteros como float; 46.0 debe guardarse como 46.
    encabezados = HEADER.strip().split(",")
    fila = [
        "MEDICINA GENERAL",
        46.0,
        "FEMENINO",
        "RESPIRATORIO ADULTO",
        "ALTA",
        "CANCER PULMONAR",
        "Paciente estable",
        "",
        "CONTROL",
    ]

    response = _subir(client, "edad_float.xlsx", _xlsx([encabezados, fila]))

    assert response.status_code == 200
    assert response.json()["inserted"] == 1


def test_xlsx_saltea_las_filas_vacias(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    encabezados = HEADER.strip().split(",")
    fila = [
        "MEDICINA GENERAL",
        46,
        "FEMENINO",
        "RESPIRATORIO ADULTO",
        "ALTA",
        "CANCER PULMONAR",
        "Paciente estable",
        "",
        "CONTROL",
    ]

    response = _subir(
        client,
        "con_vacias.xlsx",
        _xlsx([encabezados, fila, [None] * 9, fila]),
    )

    assert response.status_code == 200
    assert response.json()["inserted"] == 2


# --------------------------------------------------------------------------
# Validacion de contenido
# --------------------------------------------------------------------------


def test_archivo_solo_con_encabezados_devuelve_400(client: TestClient) -> None:
    response = _subir(client, "solo_header.csv", HEADER)

    assert response.status_code == 400
    assert response.json()["detail"] == "El archivo no contiene filas de datos"


def test_faltan_encabezados_obligatorios_devuelve_400(client: TestClient) -> None:
    contenido = "ESPEC_ORIGEN,EDAD,SEXO\nMEDICINA GENERAL,46,FEMENINO\n"

    response = _subir(client, "incompleto.csv", contenido)

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert "Faltan encabezados obligatorios" in detail
    assert "ESPEC_DESTINO" in detail
    assert "MOTIVO_INTERCONSULTA" in detail


def test_edad_no_numerica_rechaza_la_fila(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    contenido = HEADER + FILA + FILA.replace(",46,", ",cuarenta y seis,")

    response = _subir(client, "edad_texto.csv", contenido)

    assert response.status_code == 200
    body = response.json()
    assert body["inserted"] == 1
    assert body["rejected_count"] == 1
    assert body["rejected"][0]["fila"] == 3
    assert body["rejected"][0]["campos_faltantes"] == ["EDAD (formato inválido)"]


def test_edad_con_espacios_intercalados_se_rechaza(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    """ "4 6" se leia como 46. Es mas probable que sea un error de tipeo."""
    contenido = HEADER + FILA.replace(",46,", ',"4 6",')

    response = _subir(client, "edad_espacios.csv", contenido)

    assert response.status_code == 200
    body = response.json()
    assert body["inserted"] == 0
    assert body["rejected"][0]["campos_faltantes"] == ["EDAD (formato inválido)"]


def test_fila_sin_campo_obligatorio_se_rechaza_con_el_detalle(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    contenido = HEADER + FILA + FILA.replace("CANCER PULMONAR", "")

    response = _subir(client, "sin_historia.csv", contenido)

    assert response.status_code == 200
    body = response.json()
    assert body["rejected_count"] == 1
    assert body["rejected"][0]["campos_faltantes"] == ["HISTORIA_CLINICA"]


# --------------------------------------------------------------------------
# Fecha de emision (HU3-c1)
# --------------------------------------------------------------------------


def test_fecha_emision_del_archivo_se_persiste(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    contenido = HEADER.replace("\n", ",FECHA_EMISION\n") + FILA.replace(
        "\n", ",15/03/2026\n"
    )

    response = _subir(client, "con_fecha.csv", contenido)

    assert response.status_code == 200
    interconsulta_id = response.json()["ids"][0]

    detalle = client.get(f"/api/interconsultas/{interconsulta_id}")
    assert detalle.status_code == 200
    assert detalle.json()["fecha_emision"].startswith("2026-03-15")


def test_fecha_emision_no_reconocida_no_rompe_la_carga(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    contenido = HEADER.replace("\n", ",FECHA_EMISION\n") + FILA.replace(
        "\n", ",15.03.2026\n"
    )

    response = _subir(client, "fecha_rara.csv", contenido)

    assert response.status_code == 200
    interconsulta_id = response.json()["ids"][0]

    detalle = client.get(f"/api/interconsultas/{interconsulta_id}")
    assert detalle.json()["fecha_emision"] is None


# --------------------------------------------------------------------------
# Fallo de base de datos
# --------------------------------------------------------------------------


class SessionQueFalla:
    def __init__(self) -> None:
        self.rollbacks = 0
        self.commits = 0
        self.cerrada = False

    def execute(self, sql: object, params: object = None) -> None:
        raise SQLAlchemyError("conexion perdida")

    def commit(self) -> None:
        self.commits += 1

    def rollback(self) -> None:
        self.rollbacks += 1

    def close(self) -> None:
        self.cerrada = True


def test_error_de_base_de_datos_hace_rollback_y_devuelve_500(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = SessionQueFalla()
    monkeypatch.setattr(main_module, "SessionLocal", lambda: session)

    response = _subir(client, "ok.csv", HEADER + FILA)

    assert response.status_code == 500
    assert "Error en la base de datos" in response.json()["detail"]
    assert session.rollbacks == 1
    assert session.commits == 0
    assert session.cerrada is True


def test_extension_no_soportada_devuelve_400(client: TestClient) -> None:
    response = _subir(client, "datos.txt", HEADER + FILA)

    assert response.status_code == 400
    assert response.json()["detail"] == "El archivo debe ser CSV o XLSX valido"


def test_archivo_sin_nombre_devuelve_400(client: TestClient) -> None:
    response = client.post("/upload-csv", files={"file": ("", b"contenido")})

    assert response.status_code in (400, 422)


class SessionQueLanzaHttp:
    def __init__(self) -> None:
        self.rollbacks = 0
        self.cerrada = False

    def execute(self, sql: object, params: object = None) -> None:
        raise HTTPException(status_code=409, detail="conflicto de insercion")

    def commit(self) -> None:
        return None

    def rollback(self) -> None:
        self.rollbacks += 1

    def close(self) -> None:
        self.cerrada = True


def test_httpexception_durante_la_insercion_hace_rollback_y_se_propaga(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    # La HTTPException no se convierte en 500: se re-lanza tal cual, pero la
    # transaccion igual se revierte antes.
    session = SessionQueLanzaHttp()
    monkeypatch.setattr(main_module, "SessionLocal", lambda: session)

    response = _subir(client, "ok.csv", HEADER + FILA)

    assert response.status_code == 409
    assert response.json()["detail"] == "conflicto de insercion"
    assert session.rollbacks == 1
    assert session.cerrada is True


def test_edad_decimal_no_se_convierte_en_otro_numero(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    """Regresion: "53.0" -lo que exporta cualquier planilla- se guardaba como
    530 anios, y de ahi pasaba al texto que ve el modelo y al export."""
    contenido = HEADER + FILA.replace(",46,", ",53.0,")

    response = _subir(client, "edad_decimal.csv", contenido)

    assert response.status_code == 200
    assert response.json()["rejected_count"] == 0

    listado = client.get("/api/interconsultas").json()
    assert [interconsulta["edad"] for interconsulta in listado] == [53]


def test_edad_fuera_de_rango_se_rechaza_con_su_propio_motivo(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    contenido = HEADER + FILA + FILA.replace(",46,", ",530,")

    response = _subir(client, "edad_imposible.csv", contenido)

    assert response.status_code == 200
    body = response.json()
    assert body["inserted"] == 1
    assert body["rejected_count"] == 1
    assert body["rejected"][0]["campos_faltantes"] == ["EDAD (fuera de rango)"]


def test_edad_cero_se_acepta(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    """Un recien nacido es una edad valida, no un campo vacio."""
    contenido = HEADER + FILA.replace(",46,", ",0,")

    response = _subir(client, "edad_cero.csv", contenido)

    assert response.status_code == 200
    assert response.json()["rejected_count"] == 0

    listado = client.get("/api/interconsultas").json()
    assert [interconsulta["edad"] for interconsulta in listado] == [0]


def test_la_fila_rechazada_conserva_su_numero_de_linea_real(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    """Las filas en blanco se saltean al leer: numerarlas despues corria el
    indice y el modal de errores senalaba una fila distinta a la del archivo."""
    contenido = HEADER + FILA + BLANCO + VACIA + FILA.replace(",46,", ",,")

    response = _subir(client, "con_blancos.csv", contenido)

    assert response.status_code == 200
    body = response.json()
    assert body["inserted"] == 1
    assert body["rejected_count"] == 1
    # linea 1 encabezado, 2 valida, 3 y 4 en blanco, 5 la incompleta
    assert body["rejected"][0]["fila"] == 5


def test_una_fila_con_menos_columnas_se_rechaza_sola(
    client: TestClient,
    ingesta_con_modelo: object,
) -> None:
    contenido = HEADER + FILA + CORTA

    response = _subir(client, "corta.csv", contenido)

    assert response.status_code == 200
    body = response.json()
    assert body["inserted"] == 1
    assert body["rejected_count"] == 1
    assert "llegaron 2" in body["rejected"][0]["campos_faltantes"][0]


class HojaIrregular:
    """Filas de distinto largo.

    openpyxl rellena todas las filas hasta la columna mas larga, asi que un
    archivo escrito con la propia libreria nunca llega con celdas de mas: la
    rama existe para libros generados por otras herramientas.
    """

    def __init__(self, filas: list[tuple[object, ...]]) -> None:
        self._filas = filas

    def iter_rows(self, values_only: bool = False) -> object:
        return iter(self._filas)


class LibroIrregular:
    def __init__(self, filas: list[tuple[object, ...]]) -> None:
        self.active = HojaIrregular(filas)


def test_xlsx_con_celdas_de_mas_rechaza_solo_esa_fila(
    client: TestClient,
    ingesta_con_modelo: object,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Una fila con datos fuera del encabezado no tumba el archivo entero."""
    filas = [
        tuple(ENCABEZADOS_XLSX),
        tuple(FILA_XLSX),
        (*FILA_XLSX, "sobra"),
    ]
    monkeypatch.setattr(
        main_module, "load_workbook", lambda *_, **__: LibroIrregular(filas)
    )

    response = client.post(
        "/upload-csv",
        files={
            "file": (
                "extra.xlsx",
                b"no importa: load_workbook esta sustituido",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["inserted"] == 1
    assert body["rejected_count"] == 1
    assert body["rejected"][0]["fila"] == 3
    assert body["rejected"][0]["campos_faltantes"] == [
        "Contiene columnas extra con datos"
    ]


def test_si_el_modelo_ner_no_carga_la_carga_igual_termina(
    client: TestClient,
    ingesta_con_modelo: object,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """El NER es accesorio: que no cargue no puede impedir guardar el archivo."""

    def explota() -> object:
        raise OSError("modelo NER no disponible")

    monkeypatch.setattr(main_module, "get_extractor_ner", explota)

    response = _subir(client, "sin_ner.csv", HEADER + FILA)

    assert response.status_code == 200
    assert response.json()["inserted"] == 1

    listado = client.get("/api/interconsultas").json()
    assert "modelo NER" in listado[0]["entidades_error"]
